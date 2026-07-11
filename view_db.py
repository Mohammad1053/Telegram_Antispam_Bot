"""
Database Managing Script
Run this script beside the bot.py!

Choose one of this numbers:
    1 -> Print Database in excel file
    2 -> Database report in terminal
    3 -> Reset database
"""

import sqlite3
from pathlib import Path

DB_PATH = Path(__file__).resolve().parent / "antispam.db"
EXPORT_PATH = Path(__file__).resolve().parent / "antispam_export.xlsx"

USERS_COLUMNS = [
    "user_id",
    "first_name",
    "first_seen",
    "message_count",
    "spam_count",
    "is_whitelisted",
]
FLAGGED_COLUMNS = [
    "id",
    "user_id",
    "chat_id",
    "chat_title",
    "message_text",
    "score",
    "reasons",
    "action",
    "timestamp",
]


def get_connection() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


# ---------------------------------------------------------------------------
# Section 2: database report
# ---------------------------------------------------------------------------


def _print_table(cur, table_name, columns):
    print(f"\n{'=' * 60}")
    print(f"  جدول: {table_name}")
    print(f"{'=' * 60}")

    cur.execute(f"SELECT * FROM {table_name}")
    rows = cur.fetchall()

    if not rows:
        print("  (خالیه - هنوز رکوردی ثبت نشده)")
        return

    for row in rows:
        print("-" * 60)
        for col in columns:
            print(f"  {col}: {row[col]}")


def show_report():
    if not DB_PATH.exists():
        print(f"فایل دیتابیس پیدا نشد: {DB_PATH}")
        print("احتمالاً هنوز ربات رو اجرا نکردی یا هنوز پیامی رد و بدل نشده.")
        return

    conn = get_connection()
    cur = conn.cursor()

    _print_table(cur, "users", USERS_COLUMNS)
    _print_table(cur, "flagged_messages", FLAGGED_COLUMNS)

    conn.close()
    print(f"\n{'=' * 60}\n")


# ---------------------------------------------------------------------------
# Section 1: Print in excell
# ---------------------------------------------------------------------------


def _style_workbook(path: Path):
    """یه استایل ساده و حرفه‌ای (فونت، هدر بولد، عرض ستون مناسب) به فایل اضافه می‌کنه."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font

    wb = load_workbook(path)

    for sheet in wb.worksheets:
        # هدر: بولد
        for cell in sheet[1]:
            cell.font = Font(name="Arial", bold=True)

        # بقیه‌ی سلول‌ها: فونت معمولی
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Arial")

        # عرض ستون‌ها متناسب با محتوا
        for column_cells in sheet.columns:
            length = max(
                (len(str(cell.value)) if cell.value is not None else 0)
                for cell in column_cells
            )
            col_letter = column_cells[0].column_letter
            sheet.column_dimensions[col_letter].width = min(max(length + 2, 10), 50)

    wb.save(path)


def export_to_excel():
    if not DB_PATH.exists():
        print(f"فایل دیتابیس پیدا نشد: {DB_PATH}")
        print("احتمالاً هنوز ربات رو اجرا نکردی یا هنوز پیامی رد و بدل نشده.")
        return

    import pandas as pd

    conn = get_connection()
    users_df = pd.read_sql_query("SELECT * FROM users", conn)
    flagged_df = pd.read_sql_query("SELECT * FROM flagged_messages", conn)
    conn.close()

    with pd.ExcelWriter(EXPORT_PATH, engine="openpyxl") as writer:
        users_df.to_excel(writer, sheet_name="Users", index=False)
        flagged_df.to_excel(writer, sheet_name="Flagged Messages", index=False)

    _style_workbook(EXPORT_PATH)

    print(f"✅ فایل اکسل ساخته شد: {EXPORT_PATH}")


# ---------------------------------------------------------------------------
# Section 3: Database reset
# ---------------------------------------------------------------------------


def reset_database():
    if not DB_PATH.exists():
        print("دیتابیس اصلاً وجود نداره؛ چیزی برای ریست‌کردن نیست.")
        return

    confirm = input(
        "⚠️  این کار همه‌ی داده‌های دیتابیس (کاربران و پیام‌های ثبت‌شده) رو "
        "برای همیشه پاک می‌کنه و قابل بازگشت نیست.\n"
        "برای تأیید بنویس 'بله' و اینتر بزن (هر چیز دیگه‌ای لغو می‌شه): "
    ).strip()

    if confirm != "بله":
        print("لغو شد. هیچ تغییری اعمال نشد.")
        return

    conn = get_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM users")
    cur.execute("DELETE FROM flagged_messages")

    # ریست شمارنده‌ی AUTOINCREMENT جدول flagged_messages (اگه وجود داشته باشه)
    try:
        cur.execute("DELETE FROM sqlite_sequence WHERE name='flagged_messages'")
    except sqlite3.OperationalError:
        pass  # یعنی هنوز هیچ رکوردی درج نشده بود که این جدول ساخته بشه

    conn.commit()
    conn.close()

    print("✅ دیتابیس با موفقیت ریست شد و کاملاً خالیه.")


# ---------------------------------------------------------------------------
# Main menu
# ---------------------------------------------------------------------------


def main():
    print("چیکار می‌خوای بکنی؟\n")
    print("  1) خروجی‌گرفتن از دیتابیس به فایل اکسل")
    print("  2) دیدن گزارش دیتابیس (متنی، توی همین ترمینال)")
    print("  3) ریست کامل دیتابیس (پاک‌کردن همه‌ی داده‌ها)")

    choice = input("\nعدد گزینه‌ی موردنظر رو وارد کن: ").strip()

    if choice == "1":
        export_to_excel()
    elif choice == "2":
        show_report()
    elif choice == "3":
        reset_database()
    else:
        print("گزینه‌ی نامعتبر. باید 1، 2 یا 3 وارد کنی.")


if __name__ == "__main__":
    main()
